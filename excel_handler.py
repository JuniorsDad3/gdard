import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from flask import current_app

class ExcelDatabase:
    def __init__(self):
        self.file_path = current_app.config['EXCEL_DB_PATH']
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        Path(self.file_path).parent.mkdir(parents=True, exist_ok=True)
        if not Path(self.file_path).exists():
            # Create initial Excel file with required sheets
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                pd.DataFrame(columns=[
                    'id', 'username', 'password', 'role', 'farm_id', 'district'
                ]).to_excel(writer, sheet_name='users', index=False)
                pd.DataFrame(columns=[
                    'id', 'name', 'location', 'size', 'main_crop', 'livestock', 'owner_id',
                    'soil_type', 'irrigation', 'sensors_installed'
                ]).to_excel(writer, sheet_name='farms', index=False)
                pd.DataFrame(columns=[
                    'id', 'farm_id', 'date', 'report_type', 'data', 'approved'
                ]).to_excel(writer, sheet_name='reports', index=False)
                pd.DataFrame(columns=[
                    'id', 'sensor_id', 'farm_id', 'timestamp', 'temperature', 
                    'humidity', 'soil_moisture', 'light_intensity'
                ]).to_excel(writer, sheet_name='sensor_data', index=False)
    
    def get_users(self):
        return pd.read_excel(self.file_path, sheet_name='users')
    
    def add_user(self, user_data):
        wb = load_workbook(self.file_path)
        ws = wb['users']
        new_row = [
            len(ws['A']),  # ID
            user_data['username'],
            user_data['password'],
            user_data['role'],
            user_data.get('farm_id', ''),
            user_data.get('district', '')
        ]
        ws.append(new_row)
        wb.save(self.file_path)
    
    # Similar methods for other sheets (farms, reports, sensor_data)